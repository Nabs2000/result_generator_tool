# result_generator.py
from pathlib import Path
from typing import List

import datetime

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment
from openpyxl.comments import Comment

from resultsparser.student.student import Student
class ResultsGenerator:

    def __init__(self, student_list: List[Student]):
        self.student_list = student_list
        self.student_data = {}
    def output_results(self):
        example_student = self.student_list[0]
        example_res_dict = example_student.res_dict
        self.student_data["Full Name"] = []
        self.student_data["Email address"] = []
        for col_name in example_res_dict:
            self.student_data[col_name] = []

        for student in self.student_list:
            self.student_data["Full Name"].append(student.student_full_name)
            self.student_data["Email address"].append(student.student_email)
            for col_name in student.res_dict:
                self.student_data[col_name].append(student.res_dict[col_name])

        # Now, we have the dictionary, so we convert it to DF

        updated_student_data_df = pd.DataFrame.from_dict(data=self.student_data)
        desired_cols = [col for col in updated_student_data_df.columns if
                        'bonus' not in col.lower() and 'retest' not in col.lower()]
        desired_student_data = updated_student_data_df[desired_cols]
        desired_student_data_sorted = desired_student_data.sort_values(by="Full Name")
        curr_time = datetime.datetime.now()
        fmt_curr_time = curr_time.strftime("%d-%b-%Y-%H:%M:%S")
        if Path("output").exists():
            for file in Path("output").glob("*.xlsx"):
                file.unlink()
            Path("output").rmdir()
        Path("output").mkdir()
        file_name = Path(f"output/outputResultToolGenFile {fmt_curr_time}.xlsx")
        desired_student_data_sorted.to_excel(file_name, engine="openpyxl", index=False)
        workbook = load_workbook(filename=file_name)
        self.add_highlights_and_comments(wb=workbook)
        self.wrap_col_names(wb=workbook)
        workbook.save(file_name)

    def add_highlights_and_comments(self, wb):
        # Select the active sheet
        ws = wb.active
        # Mapping the column names to excel column numbers
        headings = {}
        for row in ws.iter_rows(values_only=True, min_row=1, max_row=1):
            counter = 1
            for header in row:
                headings[header] = counter
                counter += 1

        # Access cell values
        for row_num, row in enumerate(ws.iter_rows(values_only=True)):
            for student in self.student_list:
                student_name_in_spreadsheet = row[0]
                student_comment = student.changed_quizzes
                if student_name_in_spreadsheet == student.student_full_name and student_comment:
                    col_names = student_comment.keys()
                    # print(f"Student Name: {student.student_full_name}\n"
                    #       f"Column Names: {col_names}\n")
                    for col_name in col_names:
                        cell = ws.cell(row=row_num + 1, column=headings[col_name])
                        # print(f"Row: {row_num + 1}\n"
                        #       f"Column: {headings[col_name]}\n\n")
                        cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                        comment = Comment(text=student_comment[col_name], author="Alhuda Results Tool")
                        cell.comment = comment
    def wrap_col_names(self, wb):
        ws = wb.active
        for col in ws.columns:
            max_length = 0
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)/4
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2  # Adjusted width with padding
            ws.column_dimensions[col[0].column_letter].width = adjusted_width
        heading_row = list(ws.iter_rows())[0]
        for col_num, col in enumerate(heading_row):
            cell = ws.cell(row=1, column=col_num+1)
            cell.alignment = Alignment(wrapText=True)